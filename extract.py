import re
import openpyxl
#Openpyxl is a popular Python library designed for reading and writing data to Excel files (.xlsx format).
import pandas as pd
def extract_drug_info(text):
  drug_names = []
  durations = []
  timings = []


  duration_pattern = r'(\d+\s?(?:day|days|week|weeks|month|months))\b'
  timing_pattern = r'(\d+\s?(?:morning|evening|night|daily))\b'
  drug_name_pattern = r'\b(?:(?:TAB|CAP|TABLET|INJECTION|CREAM)\s+[A-Za-z][a-z0-9/]+|(?:[A-Za-z][a-z0-9/]+\s+(?:TAB|CAP|TABLET|INJECTION|CREAM)))\b'

  for line in text.splitlines():
    # find all non-overlapping matches of a regex pattern within a string.
    drug_name_match = re.findall(drug_name_pattern, line, re.IGNORECASE)
    duration_match = re.findall(duration_pattern, line, re.IGNORECASE)
    timing_match = re.findall(timing_pattern, line, re.IGNORECASE)

    if drug_name_match:
      drug_names.append(drug_name_match[0].strip())
      durations.append(duration_match[0] if duration_match else None)
      timings.append(timing_match[0] if timing_match else None)

  return drug_names, durations, timings


def write_to_excel(drug_names,  durations, timings, filename="drug_info.xlsx"):
    """
    Writes extracted drug information to an Excel file.
    """
    #create a new Excel workbook and to access the active worksheet within that workbook.
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "Medicine Name"
    sheet.cell(row=1, column=2).value = "Duration"
    sheet.cell(row=1, column=3).value = "Time"
#The enumerate function in Python is a built-in function used to iterate over iterables (like lists, tuples, strings)
# and keep track of the index (position) for each element.
    for i, (name, duration, timing) in enumerate(zip(drug_names, durations, timings), start=2):
        sheet.cell(row=i, column=1).value = name
        sheet.cell(row=i, column=2).value = duration
        sheet.cell(row=i, column=3).value = timing

    workbook.save(filename)

def process_text(text):
    drug_names, durations, timings = extract_drug_info(text)
    write_to_excel(drug_names, durations, timings)
    print("Drug information written to 'drug_info.xlsx'.")

    # Read the two Excel files into DataFrames
    df1 = pd.read_excel('drug_info.xlsx')
    df2 = pd.read_excel('medicine_data.xlsx')

    # Merge the DataFrames based on the common column 'Medicine Name'
    merged_df = pd.merge(df1, df2, on='Medicine Name', how='outer')
    merged_df['Time_x'] = merged_df['Time_x'].replace({'1 Morning': '9:00', '1 Night': '21:00', '1 Afternoon': '15:00'})

    # Write the merged DataFrame to a new Excel file
    merged_df.to_excel('merged_file.xlsx', index=False)

    print("Merged file saved as 'merged_file.xlsx'.")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        text = sys.argv[1]
        process_text(text)
    else:
        print("No text input provided.")